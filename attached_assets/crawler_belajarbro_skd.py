"""
═══════════════════════════════════════════════════════════════════════════════
  CRAWLER SOAL SKD CPNS - BELAJARBRO.ID (TWK + TIU + TKP)
═══════════════════════════════════════════════════════════════════════════════

Fungsi: Mengambil semua soal TWK, TIU, dan TKP dari belajarbro.id
         dan mengkompilasi ke 3 file Excel terpisah (1 per jenis tes).

Sumber:
  - https://belajarbro.id/cpns/soal-twk-skd
  - https://belajarbro.id/cpns/soal-tiu-skd
  - https://belajarbro.id/cpns/soal-tkp-skd

CARA PAKAI:
  1. Install dependencies:
       pip install requests beautifulsoup4 openpyxl lxml

  2. Jalankan script:
       python crawler_belajarbro_skd.py

  3. Pilih mode di awal:
       - all  : crawl semua (TWK + TIU + TKP) — paling lama, ~2-3 jam
       - twk  : hanya TWK (~30-45 menit)
       - tiu  : hanya TIU (~30-45 menit)
       - tkp  : hanya TKP (~15-20 menit)

  4. Tunggu hingga selesai. Hasilnya 1 file Excel per jenis tes.

ETIKA CRAWLING:
  - Delay 3 detik antar request (sopan ke server)
  - User-Agent jujur (mengidentifikasi sebagai script edukasi pribadi)
  - Cache otomatis: kalau script crash, jalankan ulang → halaman yang
    sudah di-download tidak akan diunduh lagi.
  - Hasil HANYA untuk belajar pribadi. Konten © Belajarbro.id.

═══════════════════════════════════════════════════════════════════════════════
"""

import os
import re
import sys
import time
import hashlib
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# KONFIGURASI UTAMA
# ─────────────────────────────────────────────────────────────────────────────

BASE_URL = "https://belajarbro.id"
CACHE_DIR = Path("cache_belajarbro")
DELAY_SECONDS = 3   # jeda antar request (jangan diturunkan di bawah 2!)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; PersonalStudyScraper/1.0; "
                  "tujuan: latihan CPNS pribadi)",
    "Accept-Language": "id-ID,id;q=0.9,en;q=0.8",
}

# Konfigurasi per jenis tes. Tiap entri:
#   - index_url   : halaman daftar paket
#   - output_file : nama file Excel hasil
#   - patterns    : dict {nama_kategori: pattern_url}
#                   pattern dipakai untuk filter link mana yang termasuk
#                   kategori tersebut.
KONFIG = {
    "TWK": {
        "index_url": f"{BASE_URL}/cpns/soal-twk-skd",
        "output_file": "Soal_TWK_Belajarbro.xlsx",
        "patterns": {
            "Nasionalisme":      "/cpns/skd/twk/soal-nasionalisme-",
            "Integritas":        "/cpns/skd/twk/soal-integritas-",
            "Bela Negara":       "/cpns/skd/twk/soal-bela-negara-",
            "Pilar Negara":      "/cpns/skd/twk/soal-pilar-negara-",
            "Bahasa Indonesia":  "/cpns/skd/twk/soal-bahasa-indonesia-",
            "Campuran":          "/cpns/soal/twk/",
        },
    },
    "TIU": {
        "index_url": f"{BASE_URL}/cpns/soal-tiu-skd",
        "output_file": "Soal_TIU_Belajarbro.xlsx",
        "patterns": {
            "Analogi Kata":       "/cpns/skd/tiu/soal-analogi-kata-",
            "Silogisme":          "/cpns/skd/tiu/soal-silogisme-",
            "Deret Angka":        "/cpns/skd/tiu/soal-deret-angka-",
            "Figural Analogi":    "/figural/soal-analogi-gambar-",
            "Figural Ketidaksamaan": "/figural/soal-ketidaksamaan-gambar-",
            "Figural Serial":     "/figural/soal-serial-gambar-",
            "Campuran":           "/cpns/soal/tiu/",
        },
    },
    "TKP": {
        "index_url": f"{BASE_URL}/cpns/soal-tkp-skd",
        "output_file": "Soal_TKP_Belajarbro.xlsx",
        "patterns": {
            "Pelayanan Publik":   "/cpns/skd/tkp/soal-pelayanan-publik-",
            "Jejaring Kerja":     "/cpns/skd/tkp/soal-jejaring-kerja-",
            "Sosial Budaya":      "/cpns/skd/tkp/soal-sosial-budaya-",
            "TIK":                "/cpns/skd/tkp/soal-teknologi-informasi-dan-komunikasi-",
            "Profesionalisme":    "/cpns/skd/tkp/soal-profesionalisme-",
            "Anti Radikalisme":   "/cpns/skd/tkp/soal-anti-radikalisme-",
            "Campuran":           "/cpns/soal/tkp/",
        },
    },
}

# Daftar URL yang DIKECUALIKAN (halaman "Semua Paket" yang bukan paket nyata)
URL_BLACKLIST = [
    "/cpns/soal-tiu-analogi-kata",
    "/cpns/soal-tiu-silogisme",
    "/cpns/soal-tiu-deret-angka",
    "/cpns/soal-tiu-figural",
]

# ─────────────────────────────────────────────────────────────────────────────
# UTILITAS: CACHING & FETCH
# ─────────────────────────────────────────────────────────────────────────────

CACHE_DIR.mkdir(exist_ok=True)


def fetch_with_cache(url):
    """Fetch URL dengan cache lokal. Tidak download ulang halaman yang sudah ada."""
    cache_key = hashlib.md5(url.encode()).hexdigest()
    cache_file = CACHE_DIR / f"{cache_key}.html"

    if cache_file.exists():
        return cache_file.read_text(encoding="utf-8")

    print(f"  → Download: {url}")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        cache_file.write_text(resp.text, encoding="utf-8")
        time.sleep(DELAY_SECONDS)
        return resp.text
    except Exception as e:
        print(f"  ✗ Gagal fetch {url}: {e}")
        return None


def is_blacklisted(url):
    """Cek apakah URL termasuk yang harus diabaikan."""
    for blocked in URL_BLACKLIST:
        if url.rstrip("/").endswith(blocked.rstrip("/")):
            return True
    return False


def url_match_pattern(url, pattern):
    """Cek apakah URL cocok dengan pattern kategori + diakhiri nomor paket."""
    if pattern not in url:
        return False
    if is_blacklisted(url):
        return False
    # Harus ada nomor di akhir URL (boleh diikuti slash)
    sisa = url.split(pattern, 1)[1].rstrip("/")
    return sisa.isdigit()


def ambil_nomor_paket(url):
    """Ekstrak nomor paket dari URL."""
    m = re.search(r"(\d+)/?$", url.rstrip("/"))
    return int(m.group(1)) if m else 0


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: TEMUKAN SEMUA LINK PAKET
# ─────────────────────────────────────────────────────────────────────────────

def temukan_link_paket(jenis_tes):
    """Ambil daftar link paket dari halaman index untuk jenis tes tertentu."""
    konfig = KONFIG[jenis_tes]
    print(f"\n[STEP 1] Mengambil daftar paket {jenis_tes} dari halaman index...")
    html = fetch_with_cache(konfig["index_url"])
    if not html:
        return {}

    soup = BeautifulSoup(html, "lxml")
    semua_link = soup.find_all("a", href=True)

    paket_per_kategori = {nama: set() for nama in konfig["patterns"]}

    for link in semua_link:
        href = link["href"]
        if href.startswith("/"):
            href = BASE_URL + href

        for nama_kat, pattern in konfig["patterns"].items():
            if url_match_pattern(href, pattern):
                paket_per_kategori[nama_kat].add(href)

    # Konversi ke list & urutkan berdasarkan nomor paket
    hasil = {}
    for nama_kat, links_set in paket_per_kategori.items():
        hasil[nama_kat] = sorted(list(links_set), key=ambil_nomor_paket)

    print(f"\n  Paket {jenis_tes} terdeteksi dari halaman index:")
    total = 0
    for nama, links in hasil.items():
        print(f"    • {nama}: {len(links)} paket")
        total += len(links)
    print(f"  Total: {total} paket")

    return hasil


def deep_discovery(jenis_tes, paket_dict):
    """Cari paket TAMBAHAN dari pagination tiap halaman paket pertama."""
    konfig = KONFIG[jenis_tes]
    print(f"\n[STEP 1b] Deep discovery {jenis_tes}: cek pagination...")

    for nama_kat, links in list(paket_dict.items()):
        if not links:
            continue

        sample_url = links[0]
        html = fetch_with_cache(sample_url)
        if not html:
            continue

        soup = BeautifulSoup(html, "lxml")
        pattern = konfig["patterns"][nama_kat]
        tambahan = set()

        for link in soup.find_all("a", href=True):
            href = link["href"]
            if href.startswith("/"):
                href = BASE_URL + href

            if url_match_pattern(href, pattern) and href not in links:
                tambahan.add(href)

        if tambahan:
            paket_dict[nama_kat] = sorted(
                list(set(links) | tambahan), key=ambil_nomor_paket
            )

    print(f"\n  Setelah deep discovery {jenis_tes}:")
    total = 0
    for nama, links in paket_dict.items():
        print(f"    • {nama}: {len(links)} paket")
        total += len(links)
    print(f"  Total: {total} paket")
    return paket_dict


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: EKSTRAK SOAL DARI 1 HALAMAN PAKET
# ─────────────────────────────────────────────────────────────────────────────

def ekstrak_soal_dari_paket(url):
    """Parse 1 halaman paket → daftar dict soal."""
    html = fetch_with_cache(url)
    if not html:
        return []

    soup = BeautifulSoup(html, "lxml")
    soal_list = []

    main_content = soup.find("main") or soup.find("article") or soup.body
    if not main_content:
        return []

    # Strategi: cari <ol> dengan 5 item (opsi A-E) di dalam <div class="konten-soal">.
    # Halaman juga menampilkan ulang <ol> di blok pembahasan — itu harus di-skip.
    semua_ol = main_content.find_all("ol")
    for ol in semua_ol:
        parent_classes = ol.parent.get("class", []) if ol.parent else []
        if "pembahasan" in parent_classes:
            continue

        items = ol.find_all("li", recursive=False)
        if len(items) != 5:
            continue

        opsi = [it.get_text(" ", strip=True) for it in items]

        # Cari teks soal sebelum <ol>
        soal_teks = ""
        nomor = ""
        prev = ol.find_previous_sibling()
        prev_texts = []
        while prev:
            txt = prev.get_text(" ", strip=True)
            if txt:
                if re.fullmatch(r"\d+", txt):
                    nomor = txt
                    break
                else:
                    prev_texts.insert(0, txt)
            prev = prev.find_previous_sibling()

        soal_teks = " ".join(prev_texts).strip()

        # Cari pembahasan & jawaban — keduanya ada di <div class="pembahasan-wr">
        # yang muncul setelah <ol> ini (bukan sibling langsung).
        # Jawaban-nya khusus ada di child <div class="jawaban"> dengan format "Jawaban : x".
        pembahasan_teks = ""
        jawaban = ""
        pw = ol.find_next("div", class_="pembahasan-wr")
        if pw:
            jw = pw.find("div", class_="jawaban")
            if jw:
                m = re.search(r"Jawaban\s*:\s*([a-eA-E])", jw.get_text(" ", strip=True))
                if m:
                    jawaban = m.group(1).upper()

            full_text = pw.get_text(" ", strip=True)
            if jw:
                jw_text = jw.get_text(" ", strip=True)
                full_text = full_text.replace(jw_text, "", 1).strip()
            pembahasan_teks = re.sub(r"Diskusi\s*$", "", full_text).strip()
            if len(pembahasan_teks) > 1500:
                pembahasan_teks = pembahasan_teks[:1500] + "..."

        if soal_teks and len(opsi) == 5:
            soal_list.append({
                "no": nomor or str(len(soal_list) + 1),
                "soal": soal_teks,
                "a": opsi[0],
                "b": opsi[1],
                "c": opsi[2],
                "d": opsi[3],
                "e": opsi[4],
                "jawaban": jawaban or "?",
                "pembahasan": pembahasan_teks or "(tidak ditemukan)",
            })

    return soal_list


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: TULIS KE EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def tulis_excel(jenis_tes, data_per_kategori, output_path):
    """Buat file Excel: 1 sheet ringkasan + 1 sheet per kategori."""
    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    answer_fill = PatternFill("solid", start_color="C6EFCE")
    header_fill = PatternFill("solid", start_color="2E75B6")
    title_fill = PatternFill("solid", start_color="1F4E78")
    total_fill = PatternFill("solid", start_color="FFF2CC")

    # ── Sheet Ringkasan ──
    cover = wb.create_sheet("Ringkasan")
    cover["A1"] = f"KUMPULAN SOAL {jenis_tes} CPNS - BELAJARBRO.ID"
    cover["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    cover["A1"].fill = title_fill
    cover["A1"].alignment = Alignment(horizontal="center", vertical="center")
    cover.merge_cells("A1:D1")
    cover.row_dimensions[1].height = 35

    cover["A3"] = "Sumber"
    cover["B3"] = KONFIG[jenis_tes]["index_url"]
    cover["A4"] = "Tanggal Crawl"
    cover["B4"] = time.strftime("%d %B %Y, %H:%M")
    cover["A5"] = "Lisensi"
    cover["B5"] = "© Belajarbro.id - hanya untuk belajar pribadi"

    headers_cover = ["Kategori", "Paket", "Total Soal", "Sheet"]
    for col_idx, h in enumerate(headers_cover, start=1):
        c = cover.cell(row=7, column=col_idx, value=h)
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border

    row = 8
    grand_total_soal = 0
    grand_total_paket = 0
    for kategori, soal_list in data_per_kategori.items():
        paket_unik = len(set(s["sumber_paket"] for s in soal_list)) if soal_list else 0
        total_soal = len(soal_list)
        cover.cell(row=row, column=1, value=kategori).border = border
        cover.cell(row=row, column=2, value=paket_unik).border = border
        cover.cell(row=row, column=3, value=total_soal).border = border
        cover.cell(row=row, column=4, value=kategori[:30]).border = border
        grand_total_soal += total_soal
        grand_total_paket += paket_unik
        row += 1

    cover.cell(row=row, column=1, value="TOTAL")
    cover.cell(row=row, column=2, value=grand_total_paket)
    cover.cell(row=row, column=3, value=grand_total_soal)
    for col in range(1, 5):
        cover.cell(row=row, column=col).font = Font(bold=True)
        cover.cell(row=row, column=col).fill = total_fill
        cover.cell(row=row, column=col).border = border

    cover.column_dimensions["A"].width = 25
    cover.column_dimensions["B"].width = 12
    cover.column_dimensions["C"].width = 12
    cover.column_dimensions["D"].width = 25

    # ── Sheet per kategori ──
    for kategori, soal_list in data_per_kategori.items():
        if not soal_list:
            continue
        sheet_name = kategori[:30]  # max 31 char untuk nama sheet
        ws = wb.create_sheet(sheet_name)

        ws["A1"] = f"SOAL {jenis_tes} - {kategori.upper()} ({len(soal_list)} soal)"
        ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:J1")
        ws.row_dimensions[1].height = 30

        ws["A2"] = "Sumber: belajarbro.id"
        ws["A2"].font = Font(name="Arial", italic=True, size=9)
        ws.merge_cells("A2:J2")

        headers = ["No", "Paket", "Soal", "A", "B", "C", "D", "E",
                   "Jawaban", "Pembahasan"]
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=4, column=col_idx, value=h)
            c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = border

        for idx, soal in enumerate(soal_list, start=5):
            ws.cell(row=idx, column=1, value=idx - 4)
            ws.cell(row=idx, column=2, value=soal.get("sumber_paket", ""))
            ws.cell(row=idx, column=3, value=soal["soal"])
            ws.cell(row=idx, column=4, value=soal["a"])
            ws.cell(row=idx, column=5, value=soal["b"])
            ws.cell(row=idx, column=6, value=soal["c"])
            ws.cell(row=idx, column=7, value=soal["d"])
            ws.cell(row=idx, column=8, value=soal["e"])
            ws.cell(row=idx, column=9, value=soal["jawaban"])
            ws.cell(row=idx, column=10, value=soal["pembahasan"])

            for col in range(1, 11):
                cell = ws.cell(row=idx, column=col)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

            ws.cell(row=idx, column=1).alignment = Alignment(
                horizontal="center", vertical="top")
            ws.cell(row=idx, column=1).font = Font(name="Arial", size=10, bold=True)
            ws.cell(row=idx, column=9).alignment = Alignment(
                horizontal="center", vertical="center")
            ws.cell(row=idx, column=9).font = Font(
                name="Arial", size=11, bold=True, color="006100")
            ws.cell(row=idx, column=9).fill = answer_fill

        widths = [5, 10, 50, 30, 30, 30, 30, 30, 10, 50]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = w

        ws.freeze_panes = "A5"

    wb.save(output_path)
    print(f"\n✓ File Excel disimpan: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# PROSES SATU JENIS TES
# ─────────────────────────────────────────────────────────────────────────────

def proses_jenis_tes(jenis_tes):
    """Crawl satu jenis tes (TWK / TIU / TKP) lalu tulis ke Excel."""
    print("\n" + "═" * 75)
    print(f"  MEMPROSES: {jenis_tes}")
    print("═" * 75)

    paket_dict = temukan_link_paket(jenis_tes)
    paket_dict = deep_discovery(jenis_tes, paket_dict)

    total = sum(len(v) for v in paket_dict.values())
    if total == 0:
        print(f"✗ Tidak ada paket {jenis_tes} ditemukan.")
        return

    print(f"\n[STEP 2] Crawl {total} paket {jenis_tes}.")
    print(f"        Estimasi: {total * DELAY_SECONDS // 60} menit "
          f"{total * DELAY_SECONDS % 60} detik (cache di-skip)")

    data_per_kategori = {}
    for kategori, paket_list in paket_dict.items():
        print(f"\n  ── {jenis_tes} / {kategori} ({len(paket_list)} paket) ──")
        soal_kategori = []
        for i, url in enumerate(paket_list, start=1):
            paket_label = f"Paket {ambil_nomor_paket(url)}"
            print(f"    [{i}/{len(paket_list)}] {paket_label}", end="  ")

            soal = ekstrak_soal_dari_paket(url)
            for s in soal:
                s["sumber_paket"] = paket_label
            soal_kategori.extend(soal)
            print(f"→ {len(soal)} soal")

        data_per_kategori[kategori] = soal_kategori
        print(f"  Subtotal {kategori}: {len(soal_kategori)} soal")

    print(f"\n[STEP 3] Menulis Excel {jenis_tes}...")
    tulis_excel(jenis_tes, data_per_kategori, KONFIG[jenis_tes]["output_file"])

    total_soal = sum(len(v) for v in data_per_kategori.values())
    print(f"\n  ✓ {jenis_tes} selesai: {total_soal} soal di "
          f"{KONFIG[jenis_tes]['output_file']}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("═" * 75)
    print("  CRAWLER SOAL SKD CPNS - BELAJARBRO.ID")
    print("  TWK + TIU + TKP")
    print("═" * 75)

    # Mode dari argumen command-line atau input
    mode = sys.argv[1].lower() if len(sys.argv) > 1 else None

    if mode not in ("all", "twk", "tiu", "tkp"):
        print("\n  Pilih mode crawl:")
        print("    1) all  - TWK + TIU + TKP (estimasi 2-3 jam)")
        print("    2) twk  - hanya TWK")
        print("    3) tiu  - hanya TIU")
        print("    4) tkp  - hanya TKP")
        pilihan = input("\n  Ketik mode (all/twk/tiu/tkp): ").strip().lower()
        mode = pilihan if pilihan in ("all", "twk", "tiu", "tkp") else "all"

    print(f"\n  Mode dipilih: {mode.upper()}")
    print(f"  Cache folder: {CACHE_DIR}/")
    print(f"  Delay       : {DELAY_SECONDS} detik antar request")

    if mode == "all":
        for jt in ("TWK", "TIU", "TKP"):
            proses_jenis_tes(jt)
    else:
        proses_jenis_tes(mode.upper())

    print("\n" + "═" * 75)
    print("  SEMUA SELESAI! 🎯")
    print("═" * 75)
    print(f"  Cache total: {len(list(CACHE_DIR.glob('*.html')))} halaman")
    print("═" * 75)


if __name__ == "__main__":
    main()
