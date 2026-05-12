"""
═══════════════════════════════════════════════════════════════════════════════
  CRAWLER SOAL TWK CPNS - BELAJARBRO.ID
═══════════════════════════════════════════════════════════════════════════════

Fungsi: Mengambil semua soal TWK dari https://belajarbro.id/cpns/soal-twk-skd
         dan mengkompilasi ke 1 file Excel dengan 1 sheet per kategori.

Kategori yang di-crawl:
  - Nasionalisme
  - Integritas
  - Bela Negara
  - Pilar Negara
  - Bahasa Indonesia
  - Campuran TWK

CARA PAKAI:
  1. Install dependencies:
       pip install requests beautifulsoup4 openpyxl lxml

  2. Jalankan script:
       python crawler_belajarbro_twk.py

  3. Tunggu hingga selesai. Hasilnya: Soal_TWK_Belajarbro.xlsx

ETIKA CRAWLING:
  - Delay 3 detik antar request (sopan ke server)
  - User-Agent jujur (mengidentifikasi sebagai script edukasi pribadi)
  - Cache otomatis: kalau script crash di tengah jalan, jalankan ulang -
    halaman yang sudah di-download tidak akan diulang.
  - Hasil HANYA untuk belajar pribadi. Konten © Belajarbro.id.

═══════════════════════════════════════════════════════════════════════════════
"""

import os
import re
import time
import json
import hashlib
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# KONFIGURASI
# ─────────────────────────────────────────────────────────────────────────────

BASE_URL = "https://belajarbro.id"
INDEX_URL = f"{BASE_URL}/cpns/soal-twk-skd"
OUTPUT_FILE = "Soal_TWK_Belajarbro.xlsx"
CACHE_DIR = Path("cache_belajarbro")
DELAY_SECONDS = 3   # jeda antar request (jangan diturunkan di bawah 2!)

# User-Agent jujur: identifikasi sebagai script edukasi
HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; PersonalStudyScraper/1.0; "
                  "tujuan: latihan CPNS pribadi)",
    "Accept-Language": "id-ID,id;q=0.9,en;q=0.8",
}

# Kategori yang akan diproses
KATEGORI = {
    "Nasionalisme": "soal-nasionalisme",
    "Integritas": "soal-integritas",
    "Bela Negara": "soal-bela-negara",
    "Pilar Negara": "soal-pilar-negara",
    "Bahasa Indonesia": "soal-bahasa-indonesia",
    "Campuran": "soal/twk",  # path-nya berbeda untuk Campuran
}

# ─────────────────────────────────────────────────────────────────────────────
# UTILITAS: CACHING & FETCH
# ─────────────────────────────────────────────────────────────────────────────

CACHE_DIR.mkdir(exist_ok=True)


def fetch_with_cache(url):
    """Fetch URL dengan cache lokal. Tidak download ulang halaman yang sudah ada."""
    cache_key = hashlib.md5(url.encode()).hexdigest()
    cache_file = CACHE_DIR / f"{cache_key}.html"

    if cache_file.exists():
        return cache_file.read_text(encoding="utf-8")

    print(f"  → Download: {url}")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        cache_file.write_text(resp.text, encoding="utf-8")
        time.sleep(DELAY_SECONDS)  # delay setelah download (hanya saat fetch baru)
        return resp.text
    except Exception as e:
        print(f"  ✗ Gagal fetch {url}: {e}")
        return None


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: TEMUKAN SEMUA LINK PAKET DARI HALAMAN INDEX
# ─────────────────────────────────────────────────────────────────────────────

def temukan_link_paket():
    """Ambil daftar link paket dari halaman index + lakukan deep discovery."""
    print("\n[STEP 1] Mengambil daftar paket dari halaman index...")
    html = fetch_with_cache(INDEX_URL)
    if not html:
        return {}

    soup = BeautifulSoup(html, "lxml")
    semua_link = soup.find_all("a", href=True)

    paket_per_kategori = {nama: [] for nama in KATEGORI}

    for link in semua_link:
        href = link["href"]
        # Normalisasi URL
        if href.startswith("/"):
            href = BASE_URL + href

        for nama_kat, pattern in KATEGORI.items():
            if pattern in href and href not in paket_per_kategori[nama_kat]:
                paket_per_kategori[nama_kat].append(href)

    # Cetak ringkasan
    print("\n  Paket terdeteksi dari halaman index:")
    total = 0
    for nama, links in paket_per_kategori.items():
        print(f"    • {nama}: {len(links)} paket")
        total += len(links)
    print(f"  Total: {total} paket\n")

    return paket_per_kategori


def deep_discovery(paket_dict):
    """Cari paket TAMBAHAN dari pagination tiap halaman paket."""
    print("[STEP 1b] Deep discovery: cek pagination di tiap kategori...")

    for nama_kat, links in paket_dict.items():
        if not links:
            continue

        # Cek paket pertama untuk lihat pagination
        sample_url = links[0]
        html = fetch_with_cache(sample_url)
        if not html:
            continue

        soup = BeautifulSoup(html, "lxml")
        # Cari link pagination yang sesuai pola kategori
        for link in soup.find_all("a", href=True):
            href = link["href"]
            if href.startswith("/"):
                href = BASE_URL + href

            pattern = KATEGORI[nama_kat]
            if pattern in href and href not in paket_dict[nama_kat]:
                # Validasi: URL harus diakhiri angka (paket-N)
                if re.search(r"-(\d+)/?$", href) or re.search(r"/(\d+)/?$", href):
                    paket_dict[nama_kat].append(href)

    # Urutkan berdasarkan nomor paket
    for nama_kat in paket_dict:
        paket_dict[nama_kat].sort(key=lambda u: int(
            re.search(r"(\d+)/?$", u).group(1) if re.search(r"(\d+)/?$", u) else 0
        ))

    print("\n  Setelah deep discovery:")
    total = 0
    for nama, links in paket_dict.items():
        print(f"    • {nama}: {len(links)} paket")
        total += len(links)
    print(f"  Total: {total} paket\n")
    return paket_dict


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: EKSTRAK SOAL DARI 1 HALAMAN PAKET
# ─────────────────────────────────────────────────────────────────────────────

def ekstrak_soal_dari_paket(url):
    """Parse 1 halaman paket → daftar dict soal."""
    html = fetch_with_cache(url)
    if not html:
        return []

    soup = BeautifulSoup(html, "lxml")
    soal_list = []

    # Strategi parsing: di Belajarbro.id, tiap soal dimulai dengan nomor
    # diikuti pertanyaan, lalu <ol> dengan opsi A-E, lalu "Pembahasan"
    # dengan teks "Jawaban : x"

    # Ambil teks lengkap halaman
    main_content = soup.find("main") or soup.find("article") or soup.body
    if not main_content:
        return []

    # Cari semua <ol> dengan tepat 5 item (opsi A-E)
    semua_ol = main_content.find_all("ol")
    for ol in semua_ol:
        items = ol.find_all("li", recursive=False)
        if len(items) != 5:
            continue

        opsi = [it.get_text(" ", strip=True) for it in items]

        # Cari teks soal (sebelum <ol>): biasanya di <p> sebelumnya
        # Cari nomor soal dengan cara naik ke parent dan ambil teks sebelumnya
        soal_teks = ""
        nomor = ""

        # Ambil semua siblings sebelum <ol> ini
        prev = ol.find_previous_sibling()
        prev_texts = []
        while prev:
            txt = prev.get_text(" ", strip=True)
            if txt:
                # Cek apakah ini nomor soal (angka tunggal)
                if re.fullmatch(r"\d+", txt):
                    nomor = txt
                    break
                else:
                    prev_texts.insert(0, txt)
            prev = prev.find_previous_sibling()

        soal_teks = " ".join(prev_texts).strip()

        # Cari pembahasan & jawaban setelah <ol>
        pembahasan_teks = ""
        jawaban = ""
        next_elem = ol.find_next_sibling()
        pembahasan_parts = []
        max_iter = 30  # batas keamanan
        while next_elem and max_iter > 0:
            txt = next_elem.get_text(" ", strip=True)
            # Cek apakah kita sudah masuk ke soal berikutnya
            if re.fullmatch(r"\d+", txt) and pembahasan_parts:
                break
            if txt:
                # Cek jawaban
                m = re.search(r"Jawaban\s*:\s*([a-eA-E])", txt)
                if m and not jawaban:
                    jawaban = m.group(1).upper()
                pembahasan_parts.append(txt)
            next_elem = next_elem.find_next_sibling()
            max_iter -= 1

        pembahasan_teks = " ".join(pembahasan_parts).strip()

        # Bersihkan teks pembahasan (hilangkan "Diskusi" di akhir)
        pembahasan_teks = re.sub(r"Diskusi\s*$", "", pembahasan_teks).strip()
        # Batasi panjang pembahasan
        if len(pembahasan_teks) > 1500:
            pembahasan_teks = pembahasan_teks[:1500] + "..."

        if soal_teks and len(opsi) == 5:
            soal_list.append({
                "no": nomor or str(len(soal_list) + 1),
                "soal": soal_teks,
                "a": opsi[0],
                "b": opsi[1],
                "c": opsi[2],
                "d": opsi[3],
                "e": opsi[4],
                "jawaban": jawaban or "?",
                "pembahasan": pembahasan_teks or "(tidak ditemukan)",
            })

    return soal_list


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: TULIS KE EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def tulis_excel(data_per_kategori, output_path):
    """Buat file Excel dengan 1 sheet per kategori + sheet Ringkasan."""
    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    answer_fill = PatternFill("solid", start_color="C6EFCE")
    header_fill = PatternFill("solid", start_color="2E75B6")
    title_fill = PatternFill("solid", start_color="1F4E78")

    # ── Sheet Ringkasan ──
    cover = wb.create_sheet("Ringkasan")
    cover["A1"] = "KUMPULAN SOAL TWK CPNS - BELAJARBRO.ID"
    cover["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    cover["A1"].fill = title_fill
    cover["A1"].alignment = Alignment(horizontal="center", vertical="center")
    cover.merge_cells("A1:D1")
    cover.row_dimensions[1].height = 35

    cover["A3"] = "Sumber"
    cover["B3"] = "https://belajarbro.id/cpns/soal-twk-skd"
    cover["A4"] = "Tanggal Crawl"
    cover["B4"] = time.strftime("%d %B %Y, %H:%M")
    cover["A5"] = "Lisensi Konten"
    cover["B5"] = "© Belajarbro.id - hanya untuk belajar pribadi"

    headers_cover = ["Kategori", "Paket", "Total Soal", "Sheet"]
    for col_idx, h in enumerate(headers_cover, start=1):
        c = cover.cell(row=7, column=col_idx, value=h)
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border

    row = 8
    grand_total_soal = 0
    grand_total_paket = 0
    for kategori, soal_list in data_per_kategori.items():
        paket_unik = len(set(s["sumber_paket"] for s in soal_list)) if soal_list else 0
        total_soal = len(soal_list)
        cover.cell(row=row, column=1, value=kategori).border = border
        cover.cell(row=row, column=2, value=paket_unik).border = border
        cover.cell(row=row, column=3, value=total_soal).border = border
        cover.cell(row=row, column=4, value=kategori).border = border
        grand_total_soal += total_soal
        grand_total_paket += paket_unik
        row += 1

    cover.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    cover.cell(row=row, column=2, value=grand_total_paket).font = Font(bold=True)
    cover.cell(row=row, column=3, value=grand_total_soal).font = Font(bold=True)
    for col in range(1, 5):
        cover.cell(row=row, column=col).fill = PatternFill("solid", start_color="FFF2CC")
        cover.cell(row=row, column=col).border = border

    cover.column_dimensions["A"].width = 22
    cover.column_dimensions["B"].width = 12
    cover.column_dimensions["C"].width = 12
    cover.column_dimensions["D"].width = 22

    # ── Sheet per kategori ──
    for kategori, soal_list in data_per_kategori.items():
        if not soal_list:
            continue
        ws = wb.create_sheet(kategori[:30])  # max 31 char untuk nama sheet

        # Judul
        ws["A1"] = f"SOAL TWK - {kategori.upper()} ({len(soal_list)} soal)"
        ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:J1")
        ws.row_dimensions[1].height = 30

        ws["A2"] = "Sumber: belajarbro.id"
        ws["A2"].font = Font(name="Arial", italic=True, size=9)
        ws.merge_cells("A2:J2")

        # Header
        headers = ["No", "Paket", "Soal", "A", "B", "C", "D", "E",
                   "Jawaban", "Pembahasan"]
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=4, column=col_idx, value=h)
            c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = border

        # Data
        for idx, soal in enumerate(soal_list, start=5):
            ws.cell(row=idx, column=1, value=idx - 4)
            ws.cell(row=idx, column=2, value=soal.get("sumber_paket", ""))
            ws.cell(row=idx, column=3, value=soal["soal"])
            ws.cell(row=idx, column=4, value=soal["a"])
            ws.cell(row=idx, column=5, value=soal["b"])
            ws.cell(row=idx, column=6, value=soal["c"])
            ws.cell(row=idx, column=7, value=soal["d"])
            ws.cell(row=idx, column=8, value=soal["e"])
            ws.cell(row=idx, column=9, value=soal["jawaban"])
            ws.cell(row=idx, column=10, value=soal["pembahasan"])

            for col in range(1, 11):
                cell = ws.cell(row=idx, column=col)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

            ws.cell(row=idx, column=1).alignment = Alignment(
                horizontal="center", vertical="top")
            ws.cell(row=idx, column=1).font = Font(name="Arial", size=10, bold=True)
            ws.cell(row=idx, column=9).alignment = Alignment(
                horizontal="center", vertical="center")
            ws.cell(row=idx, column=9).font = Font(
                name="Arial", size=11, bold=True, color="006100")
            ws.cell(row=idx, column=9).fill = answer_fill

        # Lebar kolom
        widths = [5, 10, 50, 30, 30, 30, 30, 30, 10, 50]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = w

        ws.freeze_panes = "A5"

    wb.save(output_path)
    print(f"\n✓ File Excel disimpan: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 75)
    print("  CRAWLER SOAL TWK CPNS - BELAJARBRO.ID")
    print("=" * 75)
    print(f"  Output       : {OUTPUT_FILE}")
    print(f"  Cache folder : {CACHE_DIR}/")
    print(f"  Delay        : {DELAY_SECONDS} detik antar request")
    print("=" * 75)

    # Step 1: temukan semua link paket
    paket_dict = temukan_link_paket()
    paket_dict = deep_discovery(paket_dict)

    # Konfirmasi sebelum lanjut
    total = sum(len(v) for v in paket_dict.values())
    if total == 0:
        print("✗ Tidak ada paket ditemukan. Periksa struktur situs.")
        return

    print(f"[STEP 2] Akan crawl {total} paket. Estimasi waktu: "
          f"{total * DELAY_SECONDS // 60} menit {total * DELAY_SECONDS % 60} detik")
    print("        (Halaman yang sudah di-cache tidak akan di-download ulang)\n")

    # Step 2: crawl tiap paket & ekstrak soal
    data_per_kategori = {}
    for kategori, paket_list in paket_dict.items():
        print(f"\n  ── Kategori: {kategori} ({len(paket_list)} paket) ──")
        soal_kategori = []
        for i, url in enumerate(paket_list, start=1):
            # Ambil nomor paket dari URL untuk label
            m = re.search(r"(\d+)/?$", url)
            paket_label = f"Paket {m.group(1)}" if m else url.split("/")[-1]
            print(f"    [{i}/{len(paket_list)}] {paket_label}", end="  ")

            soal = ekstrak_soal_dari_paket(url)
            for s in soal:
                s["sumber_paket"] = paket_label
            soal_kategori.extend(soal)
            print(f"→ {len(soal)} soal")

        data_per_kategori[kategori] = soal_kategori
        print(f"  Subtotal {kategori}: {len(soal_kategori)} soal")

    # Step 3: tulis ke Excel
    print("\n[STEP 3] Menulis ke Excel...")
    tulis_excel(data_per_kategori, OUTPUT_FILE)

    # Ringkasan akhir
    total_soal = sum(len(v) for v in data_per_kategori.values())
    print("\n" + "=" * 75)
    print("  SELESAI!")
    print("=" * 75)
    for kategori, soal_list in data_per_kategori.items():
        print(f"  • {kategori}: {len(soal_list)} soal")
    print(f"  • TOTAL: {total_soal} soal")
    print(f"\n  File output: {OUTPUT_FILE}")
    print(f"  Cache: {CACHE_DIR}/ ({len(list(CACHE_DIR.glob('*.html')))} halaman)")
    print("=" * 75)


if __name__ == "__main__":
    main()
