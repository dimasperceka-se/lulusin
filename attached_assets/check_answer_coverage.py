import openpyxl, re
from pathlib import Path

BASE = Path(r"c:\Users\dimas\Documents\Other\Lulusin\attached_assets\crawling_result")

# Multiple patterns to try, in priority order
PATTERNS = [
    r"[Jj]awaban\s*(?:yang\s*(?:paling\s*)?tepat\s*)?(?:adalah\s*)?(?:pilihan\s*)?[\(:]?\s*([a-eA-E])\b",
    r"[Jj]awabannya\s*(?:adalah\s*)?[\(:]?\s*([a-eA-E])\b",
    r"\bpilihan\s*\(?([a-eA-E])\)?\s*(?:adalah|merupakan|yaitu)",
    r"^\s*\(?([a-eA-E])\)?\s*[.:]",  # starts with letter
]

def extract(text):
    if not text or text == "(tidak ditemukan)":
        return None
    for p in PATTERNS:
        m = re.search(p, text)
        if m:
            return m.group(1).upper()
    return None

for fname in ["Soal_TWK_Belajarbro_run1.xlsx", "Soal_TWK_Belajarbro.xlsx", "Soal_TIU_Belajarbro.xlsx", "Soal_TKP_Belajarbro.xlsx"]:
    fp = BASE / fname
    if not fp.exists():
        continue
    wb = openpyxl.load_workbook(fp, read_only=True)
    total = 0
    have_explicit = 0
    have_from_pembahasan = 0
    no_answer = 0
    for s in wb.sheetnames:
        if s == "Ringkasan": continue
        ws = wb[s]
        for r in ws.iter_rows(min_row=5, values_only=True):
            if not r[2]: continue
            total += 1
            jawaban = (r[8] or "").strip().upper()
            if jawaban in "ABCDE":
                have_explicit += 1
            else:
                a = extract(r[9])
                if a:
                    have_from_pembahasan += 1
                else:
                    no_answer += 1
    print(f"{fname}: total={total}, explicit={have_explicit}, from_pembahasan={have_from_pembahasan}, no_answer={no_answer}")
